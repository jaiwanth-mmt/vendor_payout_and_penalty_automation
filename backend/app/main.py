from __future__ import annotations

from collections.abc import AsyncIterator, Iterator
from contextlib import asynccontextmanager
from datetime import datetime
from math import ceil
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook

from backend.app.core.env import load_env_file
from backend.app.core.paths import DEFAULT_END_DATE, DEFAULT_START_DATE, JOB_RUNTIME_ROOT, REPO_ROOT
from backend.app.agents.orchestrator import review_queue_row
from backend.app.agents.runner import (
    get_graph_topology,
    get_job_events,
    get_pending_interrupts,
    resume_case,
)
from backend.app.integrations.tracking import live_tracking_repository_from_env
from backend.app.models import (
    AgentCasesPageResponse,
    CategoryPreviewResponse,
    CreateJobResponse,
    FinalOutputPreviewResponse,
    GraphTopologyResponse,
    JobResponse,
    PendingInterrupt,
    ResumeCaseRequest,
    ReviewQueuePageResponse,
)
from backend.app.services.job_store import JobStore
from backend.app.services.package_builder import PACKAGE_FILENAME
from backend.app.services.pipeline import package_after_hitl, process_uploaded_workbook_async
import asyncio
import json
from backend.app.agents.portfolio import build_case_counts
from backend.app.agents.runner import build_agent_progress


RUNTIME_DIR = JOB_RUNTIME_ROOT
AGENT_PAGE_SIZE = 5
CATEGORY_PREVIEW_PAGE_SIZE = 5

load_env_file(REPO_ROOT / ".env")


@asynccontextmanager
async def lifespan(_app: FastAPI) -> AsyncIterator[None]:
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    yield


app = FastAPI(title="Agentic Loss Recovery Copilot API", version="0.2.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:4173",
        "http://127.0.0.1:4173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

job_store = JobStore()


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/api/jobs", response_model=CreateJobResponse)
async def create_job(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    start_date: str = Form(DEFAULT_START_DATE),
    end_date: str = Form(DEFAULT_END_DATE),
) -> CreateJobResponse:
    validate_date_range(start_date, end_date)
    validate_upload(file)

    job_id = uuid4().hex
    job_dir = RUNTIME_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    upload_path = job_dir / f"input{Path(file.filename or '').suffix or '.xlsx'}"
    upload_path.write_bytes(await file.read())

    job_store.create_job(
        job_id=job_id,
        original_filename=file.filename or "uploaded-workbook.xlsx",
        start_date=start_date,
        end_date=end_date,
        job_dir=job_dir,
        upload_path=upload_path,
    )
    job_store.mark_step_running(job_id, "upload_received", "Workbook uploaded")
    job_store.mark_step_completed(job_id, "upload_received", "Upload stored securely")

    background_tasks.add_task(run_processing_job, job_id, upload_path, start_date, end_date, job_dir)
    return CreateJobResponse(job_id=job_id, status="queued")


@app.get("/api/jobs/{job_id}", response_model=JobResponse)
def get_job(job_id: str) -> JobResponse:
    try:
        return job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error


@app.get("/api/jobs/{job_id}/download")
def download_job(job_id: str) -> FileResponse:
    try:
        snapshot = job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    package_path = job_store.get_package_path(job_id)
    if snapshot.status != "succeeded" or package_path is None or not package_path.exists():
        raise HTTPException(status_code=409, detail="ZIP package is not ready yet")

    filename = f"agentic_loss_recovery_{snapshot.start_date}_to_{snapshot.end_date}.zip"
    return FileResponse(package_path, media_type="application/zip", filename=filename)


@app.get("/api/jobs/{job_id}/final-output/download")
def download_final_output(job_id: str) -> FileResponse:
    try:
        snapshot = job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    final_output_path = job_store.get_final_output_path(job_id)
    if snapshot.status != "succeeded" or final_output_path is None or not final_output_path.exists():
        raise HTTPException(status_code=409, detail="Final output is not ready yet")

    return FileResponse(
        final_output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=snapshot.final_output.filename if snapshot.final_output else "final_output.xlsx",
    )


@app.get("/api/jobs/{job_id}/final-output/preview", response_model=FinalOutputPreviewResponse)
def preview_final_output(
    job_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    booking_id: str | None = Query(None),
) -> FinalOutputPreviewResponse:
    try:
        snapshot = job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    final_output_path = job_store.get_final_output_path(job_id)
    if snapshot.status != "succeeded" or final_output_path is None or not final_output_path.exists():
        raise HTTPException(status_code=409, detail="Final output is not ready yet")

    return read_final_output_preview(
        final_output_path,
        page=page,
        page_size=page_size,
        booking_id=booking_id,
    )


@app.get("/api/jobs/{job_id}/categories/{slug}/preview", response_model=CategoryPreviewResponse)
def preview_category(
    job_id: str,
    slug: str,
    page: int = Query(1, ge=1),
    booking_id: str | None = Query(None),
) -> CategoryPreviewResponse:
    try:
        snapshot = job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    if snapshot.status != "succeeded":
        raise HTTPException(status_code=409, detail="Category preview is not ready yet")

    category_path = job_store.get_category_processed_path(job_id, slug)
    if category_path is None:
        raise HTTPException(status_code=404, detail="Category not found")
    if not category_path.exists():
        raise HTTPException(status_code=409, detail="Category preview is not ready yet")

    return read_category_preview(
        category_path,
        page=page,
        page_size=CATEGORY_PREVIEW_PAGE_SIZE,
        booking_id=booking_id,
    )


@app.get("/api/jobs/{job_id}/agent-audit/download")
def download_agent_audit(job_id: str) -> FileResponse:
    try:
        snapshot = job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    path = job_store.get_agent_audit_path(job_id)
    if snapshot.status != "succeeded" or path is None or not path.exists():
        raise HTTPException(status_code=409, detail="Agent audit workbook is not ready yet")

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="agent_audit.xlsx",
    )


@app.get("/api/jobs/{job_id}/review-queue/download")
def download_review_queue(job_id: str) -> FileResponse:
    try:
        snapshot = job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    path = job_store.get_review_queue_path(job_id)
    if snapshot.status != "succeeded" or path is None or not path.exists():
        raise HTTPException(status_code=409, detail="Review queue workbook is not ready yet")

    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="review_queue.xlsx",
    )


@app.get("/api/jobs/{job_id}/review-queue", response_model=ReviewQueuePageResponse)
def list_review_queue(job_id: str, page: int = Query(1, ge=1)) -> ReviewQueuePageResponse:
    try:
        snapshot = job_store.snapshot(job_id)
        cases = job_store.get_agent_cases(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    if snapshot.status not in {"succeeded", "awaiting_review"}:
        raise HTTPException(status_code=409, detail="Review queue is not ready yet")

    items = [review_queue_row(case) for case in cases if case.get("review_status") != "auto_ready"]
    page_items, safe_page, total_pages = paginate_items(items, page=page, page_size=AGENT_PAGE_SIZE)
    return ReviewQueuePageResponse(
        items=page_items,
        item_count=len(items),
        page=safe_page,
        page_size=AGENT_PAGE_SIZE,
        total_pages=total_pages,
    )


@app.get("/api/jobs/{job_id}/cases", response_model=AgentCasesPageResponse)
def list_agent_cases(job_id: str, page: int = Query(1, ge=1)) -> AgentCasesPageResponse:
    try:
        snapshot = job_store.snapshot(job_id)
        cases = job_store.get_agent_cases(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    if snapshot.status not in {"succeeded", "awaiting_review"}:
        raise HTTPException(status_code=409, detail="Agent cases are not ready yet")

    page_cases, safe_page, total_pages = paginate_items(cases, page=page, page_size=AGENT_PAGE_SIZE)
    return AgentCasesPageResponse(
        cases=page_cases,
        case_count=len(cases),
        page=safe_page,
        page_size=AGENT_PAGE_SIZE,
        total_pages=total_pages,
    )


@app.get("/api/jobs/{job_id}/cases/{booking_id}")
def get_agent_case(job_id: str, booking_id: str) -> dict[str, Any]:
    try:
        snapshot = job_store.snapshot(job_id)
        cases = job_store.get_agent_cases(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    if snapshot.status not in {"succeeded", "awaiting_review"}:
        raise HTTPException(status_code=409, detail="Agent cases are not ready yet")

    for case in cases:
        if str(case.get("booking_id", "")).strip() == booking_id:
            return case

    raise HTTPException(status_code=404, detail="Agent case not found")


@app.get("/api/jobs/{job_id}/graph", response_model=GraphTopologyResponse)
def get_job_graph(job_id: str) -> GraphTopologyResponse:
    try:
        snapshot = job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error
    topology = snapshot.graph_topology or get_graph_topology(job_id)
    return GraphTopologyResponse(**topology)


@app.get("/api/jobs/{job_id}/interrupts")
def list_job_interrupts(job_id: str) -> list[PendingInterrupt]:
    try:
        snapshot = job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error
    return snapshot.pending_interrupts


@app.get("/api/jobs/{job_id}/events")
async def stream_job_events(job_id: str, after: int = Query(0, ge=0)) -> StreamingResponse:
    try:
        job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    async def event_generator() -> AsyncIterator[str]:
        cursor = after
        idle_rounds = 0
        while True:
            try:
                snapshot = job_store.snapshot(job_id)
            except KeyError:
                yield f"event: error\ndata: {json.dumps({'error': 'Job not found'})}\n\n"
                break

            store_events = list(snapshot.graph_events or [])
            runner_events = get_job_events(job_id, after_index=0)
            # Prefer job_store buffer; fall back to runner registry.
            events = store_events if store_events else runner_events
            if cursor < len(events):
                for event in events[cursor:]:
                    yield f"data: {json.dumps(event)}\n\n"
                cursor = len(events)
                idle_rounds = 0
            else:
                idle_rounds += 1
                yield f": keepalive {idle_rounds}\n\n"

            if snapshot.status in {"succeeded", "failed"} and idle_rounds > 2:
                yield f"event: done\ndata: {json.dumps({'status': snapshot.status})}\n\n"
                break
            await asyncio.sleep(0.5)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.post("/api/jobs/{job_id}/cases/{booking_id}/resume")
async def resume_job_case(job_id: str, booking_id: str, body: ResumeCaseRequest) -> dict[str, Any]:
    try:
        snapshot = job_store.snapshot(job_id)
    except KeyError as error:
        raise HTTPException(status_code=404, detail="Job not found") from error

    if snapshot.status not in {"awaiting_review", "running"}:
        raise HTTPException(status_code=409, detail="Job is not awaiting human review")

    human_decision = {
        "decision": body.decision,
        "review_status": body.review_status,
        "review_reason": body.review_reason,
        "rationale": body.rationale,
        "recommended_action": body.recommended_action,
    }
    if body.recommended_recovery_amount is not None:
        human_decision["recommended_recovery_amount"] = body.recommended_recovery_amount

    try:
        payload = await resume_case(
            job_id=job_id,
            booking_id=booking_id,
            human_decision=human_decision,
            on_event=lambda event: job_store.append_graph_event(job_id, event),
        )
    except KeyError as error:
        raise HTTPException(status_code=404, detail=str(error)) from error

    cases = job_store.get_agent_cases(job_id)
    updated_cases = []
    replaced = False
    for case in cases:
        if str(case.get("booking_id", "")).strip() == booking_id:
            updated_cases.append(payload)
            replaced = True
        else:
            updated_cases.append(case)
    if not replaced:
        updated_cases.append(payload)
    job_store.update_agent_cases(job_id, updated_cases)

    pending = get_pending_interrupts(job_id)
    job_store.set_pending_interrupts(job_id, pending)
    case_counts = build_case_counts(updated_cases)
    agent_progress = build_agent_progress(updated_cases)
    with job_store._lock:
        job = job_store._jobs[job_id]
        job["case_counts"] = case_counts
        job["agent_progress"] = agent_progress
        from backend.app.services.job_store import utc_now

        job["updated_at"] = utc_now()

    if not pending:
        try:
            await _complete_job_after_hitl(job_id, updated_cases)
        except Exception as error:
            job_store.fail_job(job_id, f"Packaging after human review failed: {error}")
            raise HTTPException(status_code=500, detail=str(error)) from error

    return payload


async def _complete_job_after_hitl(job_id: str, agent_cases: list[dict[str, Any]]) -> None:
    """Build portfolio + package artifacts when packaging was deferred for human review."""
    job = job_store._jobs[job_id]
    job_dir = Path(job["job_dir"])
    job_store.mark_step_running(job_id, "package_prepared", "Packaging after human review")
    packaged = await package_after_hitl(
        job_dir=job_dir,
        output_package_path=job_dir / PACKAGE_FILENAME,
        start_date=str(job.get("start_date") or ""),
        end_date=str(job.get("end_date") or ""),
        category_outputs=list(job.get("category_outputs") or []),
        agent_cases=agent_cases,
        metrics=dict(job.get("metrics") or {}),
        job_id=job_id,
        on_event=lambda event: job_store.append_graph_event(job_id, event),
    )
    job_store.mark_step_completed(job_id, "agent_investigation", "LangGraph investigation completed")
    job_store.mark_step_completed(job_id, "package_prepared", "ZIP package ready after human review")
    job_store.complete_job(
        job_id,
        metrics=packaged["metrics"],
        category_outputs=packaged["category_outputs"],
        package_path=packaged["package_path"],
        final_output_path=packaged["final_output_path"],
        final_output=packaged["final_output"],
        agent_audit_path=packaged["agent_audit_path"],
        review_queue_path=packaged["review_queue_path"],
        agent_summary_path=packaged["agent_summary_path"],
        agent_summary=packaged["agent_summary"],
        case_counts=packaged["case_counts"],
        agent_progress=packaged["agent_progress"],
        agent_cases=packaged["agent_cases"],
    )
    job_store.set_pending_interrupts(job_id, [])


async def run_processing_job(
    job_id: str,
    upload_path: Path,
    start_date: str,
    end_date: str,
    job_dir: Path,
) -> None:
    def _init_category_and_investigation(categories: list[dict[str, Any]]) -> None:
        job_store.initialize_category_progress(job_id, categories)
        total_cases = sum(int(category.get("row_count") or 0) for category in categories)
        job_store.init_investigation_progress(job_id, total_cases=total_cases)

    try:
        job_store.mark_step_running(job_id, "agent_investigation", "LangGraph investigation pending category processing")
        job_store.set_graph_topology(job_id, get_graph_topology(job_id))
        tracking_repository = live_tracking_repository_from_env()
        result = await process_uploaded_workbook_async(
            input_path=upload_path,
            tracking_repository=tracking_repository,
            output_package_path=job_dir / PACKAGE_FILENAME,
            start_date=start_date,
            end_date=end_date,
            on_step_start=lambda step_id, message: job_store.mark_step_running(job_id, step_id, message),
            on_step_complete=lambda step_id, message: job_store.mark_step_completed(job_id, step_id, message),
            on_warning=lambda warning: job_store.add_warning(job_id, warning),
            on_step_progress=lambda step_id, completed_units, total_units, message: job_store.update_step_units(
                job_id,
                step_id,
                completed_units=completed_units,
                total_units=total_units,
                message=message,
            ),
            on_category_progress_initialized=_init_category_and_investigation,
            on_category_progress=lambda slug, update: job_store.update_category_progress(job_id, slug, **update),
            job_id=job_id,
            enable_hitl=True,
            on_agent_event=lambda event: job_store.append_graph_event(job_id, event),
        )
        job_store.set_graph_topology(job_id, get_graph_topology(job_id))
        if result.awaiting_review:
            job_store.mark_awaiting_review(
                job_id,
                metrics=result.metrics,
                category_outputs=result.category_outputs,
                case_counts=result.case_counts,
                agent_progress=result.agent_progress,
                agent_cases=result.agent_cases,
                pending_interrupts=result.pending_interrupts or get_pending_interrupts(job_id),
            )
            return

        job_store.mark_step_completed(job_id, "agent_investigation", "LangGraph investigation completed")
        job_store.complete_job(
            job_id,
            metrics=result.metrics,
            category_outputs=result.category_outputs,
            package_path=result.package_path,
            final_output_path=result.final_output_path,
            final_output=result.final_output,
            agent_audit_path=result.agent_audit_path,
            review_queue_path=result.review_queue_path,
            agent_summary_path=result.agent_summary_path,
            agent_summary=result.agent_summary,
            case_counts=result.case_counts,
            agent_progress=result.agent_progress,
            agent_cases=result.agent_cases,
        )
        job_store.set_pending_interrupts(job_id, [])
    except Exception as error:
        job_store.fail_job(job_id, str(error))


def validate_iso_date(value: str, *, field_name: str) -> None:
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError as error:
        raise HTTPException(
            status_code=422,
            detail=f"{field_name} must be in YYYY-MM-DD format",
        ) from error


def validate_date_range(start_date: str, end_date: str) -> None:
    validate_iso_date(start_date, field_name="start_date")
    validate_iso_date(end_date, field_name="end_date")
    if start_date > end_date:
        raise HTTPException(status_code=422, detail="start_date must be on or before end_date")


def validate_upload(file: UploadFile) -> None:
    filename = file.filename or ""
    if Path(filename).suffix.lower() not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail="Upload a QlikSense Excel workbook (.xlsx or .xls)")


def read_final_output_preview(
    path: Path,
    *,
    page: int,
    page_size: int,
    booking_id: str | None = None,
) -> FinalOutputPreviewResponse:
    return FinalOutputPreviewResponse(
        **read_workbook_rows_page(path, page=page, page_size=page_size, booking_id=booking_id)
    )


def read_category_preview(
    path: Path,
    *,
    page: int,
    page_size: int,
    booking_id: str | None = None,
) -> CategoryPreviewResponse:
    return CategoryPreviewResponse(
        **read_workbook_rows_page(path, page=page, page_size=page_size, booking_id=booking_id)
    )


def read_workbook_rows_page(
    path: Path,
    *,
    page: int,
    page_size: int,
    booking_id: str | None = None,
) -> dict[str, Any]:
    if not path.exists() or path.stat().st_size == 0:
        raise HTTPException(status_code=404, detail="Final output workbook is not ready")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        raise HTTPException(status_code=404, detail="Final output workbook is not readable") from error
    try:
        worksheet = workbook.active
        worksheet_rows = worksheet.iter_rows(values_only=True)
        columns = [str(column or "") for column in next(worksheet_rows, ())]

        booking_filter = normalize_booking_id_filter(booking_id)
        if booking_filter:
            return read_filtered_workbook_rows_page(
                worksheet_rows,
                columns=columns,
                page=page,
                page_size=page_size,
                booking_id=booking_filter,
            )

        row_count = max((worksheet.max_row or 1) - 1, 0)
        total_pages = max(1, ceil(row_count / page_size))
        safe_page = min(page, total_pages)
        start_index = (safe_page - 1) * page_size
        end_index = start_index + page_size
        rows: list[dict[str, Any]] = []

        for row_index, values in enumerate(worksheet_rows):
            if row_index < start_index:
                continue
            if row_index >= end_index:
                break
            rows.append(serialize_preview_row(columns, values))

        return {
            "columns": columns,
            "rows": rows,
            "row_count": row_count,
            "page": safe_page,
            "page_size": page_size,
            "total_pages": total_pages,
        }
    finally:
        workbook.close()


def read_filtered_workbook_rows_page(
    worksheet_rows: Iterator[tuple[Any, ...]],
    *,
    columns: list[str],
    page: int,
    page_size: int,
    booking_id: str,
) -> dict[str, Any]:
    booking_id_column_index = find_booking_id_column_index(columns)
    matching_rows: list[dict[str, Any]] = []

    if booking_id_column_index is not None:
        for values in worksheet_rows:
            if row_booking_id(values, booking_id_column_index) == booking_id:
                matching_rows.append(serialize_preview_row(columns, values))

    row_count = len(matching_rows)
    total_pages = max(1, ceil(row_count / page_size))
    safe_page = min(page, total_pages)
    start_index = (safe_page - 1) * page_size
    end_index = start_index + page_size

    return {
        "columns": columns,
        "rows": matching_rows[start_index:end_index],
        "row_count": row_count,
        "page": safe_page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


def normalize_booking_id_filter(value: str | None) -> str:
    return str(value or "").strip()


def find_booking_id_column_index(columns: list[str]) -> int | None:
    for index, column in enumerate(columns):
        normalized_column = column.strip().lower().replace(" ", "_")
        if normalized_column == "booking_id":
            return index
    return None


def row_booking_id(values: tuple[Any, ...], column_index: int) -> str:
    if column_index >= len(values):
        return ""
    return str(values[column_index] or "").strip()


def serialize_preview_row(columns: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {
        column: serialize_preview_cell(values[column_index] if column_index < len(values) else None)
        for column_index, column in enumerate(columns)
    }


def paginate_items(items: list[Any], *, page: int, page_size: int) -> tuple[list[Any], int, int]:
    total_pages = max(1, ceil(len(items) / page_size))
    safe_page = min(page, total_pages)
    start_index = (safe_page - 1) * page_size
    end_index = start_index + page_size
    return items[start_index:end_index], safe_page, total_pages


def serialize_preview_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return value
